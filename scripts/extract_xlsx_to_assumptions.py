"""Extract historical / forecast time-series from the original xlsx into
long-format CSVs under ``assumptions/2026/timeseries/``.

CSV schema (every output file):
    country,period,scenario,value
    ZAF,2005,shared,71248.3        # historical: same value across scenarios
    ZAF,2030,high_demand,...       # forecast: scenario-specific
    ZAF,2030,low_demand,...

Run:
    python scripts/extract_xlsx_to_assumptions.py

Idempotent: re-running overwrites the CSVs from the xlsx. The xlsx itself
is treated as read-only.
"""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "docs" / "Liquid Fuels Model - Supply Demand, 2025 - Reatile Copy.xlsx"
OUT = REPO / "assumptions" / "2026" / "timeseries"
OUT.mkdir(parents=True, exist_ok=True)


def named_range_rows(wb, name: str) -> list[list]:
    """Resolve a defined name to a list-of-lists of cell values."""
    dn = wb.defined_names[name]
    rows: list[list] = []
    for sheet_title, coord in dn.destinations:
        ws = wb[sheet_title]
        cells = ws[coord]
        if hasattr(cells, "value"):
            rows.append([cells.value])
        else:
            for row in cells:
                if hasattr(row, "value"):
                    rows.append([row.value])
                else:
                    rows.append([c.value for c in row])
    return rows


def write_long_csv(path: Path, records: list[tuple]) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["country", "period", "scenario", "value"])
        w.writerows(records)
    print(f"  wrote {len(records):>5} rows -> {path.relative_to(REPO).as_posix()}")


def is_year(v) -> bool:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return 1900 <= int(v) <= 2200
    return False


def to_year(v) -> int:
    return int(v)


def extract_macro(wb) -> None:
    """GDP, GDP/capita, population — historical (shared) plus high/low forecast.

    Source: named range `GDP` at Assumptions!$Z$8:$AH$54.
    Columns: Year | GDP | GDP High | GDP Low | GDP/capita (constant 2015) |
             Population high | Population low | GDP/capita high | GDP/capita low
    """
    rows = named_range_rows(wb, "GDP")
    header = rows[0]
    assert header[0] == "Year", header
    data = [r for r in rows[1:] if is_year(r[0])]

    # Historical period: only one observed series per quantity.
    # Forecast period: scenario-specific high/low columns.
    # We treat all available columns as authoritative and emit shared rows
    # where the high/low values coincide with each other and with the base.

    gdp_records: list[tuple] = []
    gdp_pc_records: list[tuple] = []
    pop_records: list[tuple] = []

    for r in data:
        year = to_year(r[0])
        gdp_base, gdp_high, gdp_low = r[1], r[2], r[3]
        gdp_pc_base = r[4]
        pop_high, pop_low = r[5], r[6]
        gdp_pc_high, gdp_pc_low = r[7], r[8]

        # GDP — three-column source. If high == low == base, emit one shared row.
        if gdp_base == gdp_high == gdp_low and gdp_base is not None:
            gdp_records.append(("ZAF", year, "shared", gdp_base))
        else:
            if gdp_high is not None:
                gdp_records.append(("ZAF", year, "high_demand", gdp_high))
            if gdp_low is not None:
                gdp_records.append(("ZAF", year, "low_demand", gdp_low))

        # GDP/capita — historical column has the observed value;
        # high/low forecast columns hold scenario forecasts.
        if gdp_pc_base is not None and gdp_pc_high == gdp_pc_base and gdp_pc_low == gdp_pc_base:
            gdp_pc_records.append(("ZAF", year, "shared", gdp_pc_base))
        else:
            if gdp_pc_high is not None:
                gdp_pc_records.append(("ZAF", year, "high_demand", gdp_pc_high))
            if gdp_pc_low is not None:
                gdp_pc_records.append(("ZAF", year, "low_demand", gdp_pc_low))
            # Fall back: if neither scenario column has a value but the base does, treat as shared.
            if gdp_pc_high is None and gdp_pc_low is None and gdp_pc_base is not None:
                gdp_pc_records.append(("ZAF", year, "shared", gdp_pc_base))

        # Population — high/low only.
        if pop_high == pop_low and pop_high is not None:
            pop_records.append(("ZAF", year, "shared", pop_high))
        else:
            if pop_high is not None:
                pop_records.append(("ZAF", year, "high_demand", pop_high))
            if pop_low is not None:
                pop_records.append(("ZAF", year, "low_demand", pop_low))

    write_long_csv(OUT / "gdp.csv", gdp_records)
    write_long_csv(OUT / "gdp_per_capita.csv", gdp_pc_records)
    write_long_csv(OUT / "population.csv", pop_records)


def extract_gdp_growth(wb) -> None:
    """Real GDP growth paths per scenario (Assumptions sheet)."""
    high = named_range_rows(wb, "GDP_high_growth")
    low = named_range_rows(wb, "GDP_low_growth")
    records: list[tuple] = []
    for row in high[1:]:
        if is_year(row[0]) and row[1] is not None:
            records.append(("ZAF", to_year(row[0]), "high_demand", row[1]))
    for row in low[1:]:
        if is_year(row[0]) and row[1] is not None:
            records.append(("ZAF", to_year(row[0]), "low_demand", row[1]))
    write_long_csv(OUT / "gdp_growth.csv", records)


def extract_efficiency_improvement(wb) -> None:
    """Per-scenario efficiency-improvement paths for diesel and gasoline."""
    diesel_high = named_range_rows(wb, "HighEfficiency_Improvement_Diesel")
    diesel_low = named_range_rows(wb, "LowEfficiency_Improvement_Diesel")
    gasoline_high = named_range_rows(wb, "HighEfficiency_Improvement_Gasoline")
    gasoline_low = named_range_rows(wb, "LowEfficiency_Improvement_Gasoline")

    # These ranges align with GDP_high_growth's year range starting at 2020.
    growth_years = [to_year(r[0]) for r in named_range_rows(wb, "GDP_high_growth")[1:]
                    if is_year(r[0])]

    def emit(rng: list[list], scenario: str) -> list[tuple]:
        out = []
        # First row is header in some named ranges, raw value in others; skip if header.
        rows = rng[1:] if isinstance(rng[0][0], str) else rng
        for year, row in zip(growth_years, rows):
            if row[0] is not None:
                out.append(("ZAF", year, scenario, row[0]))
        return out

    diesel_records = emit(diesel_high, "high_demand") + emit(diesel_low, "low_demand")
    gasoline_records = emit(gasoline_high, "high_demand") + emit(gasoline_low, "low_demand")
    write_long_csv(OUT / "efficiency_improvement_diesel.csv", diesel_records)
    write_long_csv(OUT / "efficiency_improvement_gasoline.csv", gasoline_records)


def extract_ocgt_load_factor(wb) -> None:
    """OCGT load factor by year and scenario.

    The xlsx's `_High` / `_Low` suffix refers to grid availability (EAF), not
    OCGT utilisation. So `PowerGenLoadFactor_Low` (low EAF -> stressed grid ->
    high OCGT use) maps to OUR `high_demand` scenario, and vice versa. See the
    scenario-mapping note in `_meta.yaml`.
    """
    pessimistic_grid = named_range_rows(wb, "PowerGenLoadFactor_Low")   # high OCGT use
    optimistic_grid  = named_range_rows(wb, "PowerGenLoadFactor_High")  # low OCGT use
    records: list[tuple] = []
    for row in pessimistic_grid[1:]:
        if is_year(row[0]) and row[1] is not None:
            records.append(("ZAF", to_year(row[0]), "high_demand", row[1]))
    for row in optimistic_grid[1:]:
        if is_year(row[0]) and row[1] is not None:
            records.append(("ZAF", to_year(row[0]), "low_demand", row[1]))
    write_long_csv(OUT / "ocgt_load_factor.csv", records)


def extract_ocgt_load_shedding(wb) -> None:
    """OCGT load-shedding GWh by year and scenario. Same EAF inversion as load_factor."""
    pessimistic_grid = named_range_rows(wb, "LoadShedding_Low")   # high load shedding
    optimistic_grid  = named_range_rows(wb, "LoadShedding_High")  # low load shedding
    records: list[tuple] = []
    for row in pessimistic_grid[1:]:
        if is_year(row[0]) and row[1] is not None:
            records.append(("ZAF", to_year(row[0]), "high_demand", row[1]))
    for row in optimistic_grid[1:]:
        if is_year(row[0]) and row[1] is not None:
            records.append(("ZAF", to_year(row[0]), "low_demand", row[1]))
    write_long_csv(OUT / "ocgt_load_shedding.csv", records)


def extract_passenger_departures(wb) -> None:
    """Aviation passenger-departure series — observed history stitched with baseline forecast.

    The xlsx fitted its jet regression on observed pax. PaxBaseScenario is a
    smoothed post-pandemic baseline forecast (2020 = 21M, ignoring the actual
    COVID drop), so feeding it as input to the regression for historical years
    produces values that don't match observed jet demand.

    Stitch:
      - 2017-2024 (or whatever's filled): observed Departures_Pass from
        the `Jet - DemandSupply` sheet (column K under the un-transposed table).
      - Years beyond observed: PaxBaseScenario.
      - Observed wins for overlapping years.
    """
    # Observed from Jet-DS — un-transposed table starts at row 9 with year in col E,
    # observed pax in col K (header at row 8 says 'Departures_Pass'). The sheet has
    # several side tables further down sharing column E for years; stop the walk
    # the moment a header-like string (e.g. 'Jet Fuel', 'Year', 'Av Gas') appears.
    ws = wb["Jet - DemandSupply"]
    observed: dict[int, float] = {}
    for row_idx in range(9, 30):
        year_cell = ws.cell(row=row_idx, column=5).value
        pax_cell = ws.cell(row=row_idx, column=11).value

        # Skip blank rows mid-table; they don't end the table.
        if year_cell is None:
            continue
        # Plain numeric year, or "2023*" annotation — extract digits.
        if is_year(year_cell):
            year_int = to_year(year_cell)
        elif isinstance(year_cell, str):
            digits = "".join(c for c in year_cell if c.isdigit())
            if not digits or not (1900 <= int(digits) <= 2200):
                break  # header of next side-table — stop walking.
            year_int = int(digits)
        else:
            break
        if (
            isinstance(pax_cell, (int, float))
            and not isinstance(pax_cell, bool)
            and pax_cell > 0
        ):
            observed[year_int] = float(pax_cell)

    # Baseline forecast (post-pandemic recovery profile)
    baseline: dict[int, float] = {}
    for row in named_range_rows(wb, "PaxBaseScenario")[1:]:
        if is_year(row[0]) and row[1] is not None:
            baseline[to_year(row[0])] = float(row[1])

    # Stitch
    years = sorted(set(observed) | set(baseline))
    records: list[tuple] = []
    for year in years:
        value = observed.get(year, baseline.get(year))
        if value is None:
            continue
        records.append(("ZAF", year, "shared", value))
    write_long_csv(OUT / "passenger_departures.csv", records)


def extract_refinery_production(wb) -> None:
    """Refinery utilisation by refinery, year, and scenario.

    `Production_High` / `Production_Low` define annual utilisation rates
    per refinery; combined with refinery nameplate capacity these give
    actual production volumes.
    """
    high = named_range_rows(wb, "Production_High")
    low = named_range_rows(wb, "Production_Low")
    # Header row 1: ['Production', 'Enref', 'SAPREF', 'Natref', 'Sasol', 'Astron', 'PetroSA']
    refineries = [str(c).lower() for c in high[0][1:7]]

    records: list[tuple] = []
    for src, scenario in [(high, "high_demand"), (low, "low_demand")]:
        for row in src[1:]:
            if not is_year(row[0]):
                continue
            year = to_year(row[0])
            for refinery, util in zip(refineries, row[1:7]):
                if util is not None:
                    # Encode refinery name in the country slot — refineries are
                    # not country-keyed in this CSV. Schema bends a bit here.
                    records.append((f"ZAF:{refinery}", year, scenario, util))
    write_long_csv(OUT / "refinery_production.csv", records)


def extract_historical_demand(wb) -> None:
    """Historical RSA demand by product, 2005-onwards.

    Source: per-product DemandSupply sheets, "RSA Demand" column. These are
    the totals the original xlsx fits its product-level regressions against,
    so they're the authoritative comparator for any reconciliation.

    Output schema is `(country, period, scenario, product, value)`. Scenario is
    always "shared" since history is observed, not scenario-dependent.

    Layout notes (from xlsx inspection):
      - Gasoline - DemandSupply : Year in col C, RSA Demand in col K, header row 9
      - Diesel - DemandSupply   : Year in col C, RSA Demand in col K, header row 10
      - Jet - DemandSupply      : Year in col E, RSA Demand in col J, header row 8
    """
    sheets = [
        ("Gasoline - DemandSupply", "petrol_95",     "C", "K", 9),
        ("Diesel - DemandSupply",   "diesel_50ppm",  "C", "K", 10),
        ("Jet - DemandSupply",      "jet_a1",        "E", "J", 8),
    ]

    out_path = OUT / "historical_demand.csv"
    with out_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["country", "period", "scenario", "product", "value"])
        total = 0
        for sheet_name, product, year_col, value_col, header_row in sheets:
            ws = wb[sheet_name]
            # Walk down from the row immediately after the header until we
            # stop seeing year-like values. Stop at the first non-year row.
            for row_idx in range(header_row + 1, ws.max_row + 1):
                year_v = ws[f"{year_col}{row_idx}"].value
                if not is_year(year_v):
                    # Allow blank rows mid-table; only stop after a streak of misses.
                    if year_v is None:
                        continue
                    # Could be "2023*" annotation — try string parse.
                    if isinstance(year_v, str):
                        digits = "".join(c for c in year_v if c.isdigit())
                        if not digits or not (1900 <= int(digits) <= 2200):
                            continue
                        year_int = int(digits)
                    else:
                        continue
                else:
                    year_int = to_year(year_v)
                value = ws[f"{value_col}{row_idx}"].value
                if value is None or not isinstance(value, (int, float)) or value == 0:
                    continue
                w.writerow(["ZAF", year_int, "shared", product, float(value)])
                total += 1
    print(f"  wrote {total:>5} rows -> {out_path.relative_to(REPO).as_posix()}")


def main() -> None:
    print(f"Reading: {XLSX.relative_to(REPO).as_posix()}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    extract_macro(wb)
    extract_gdp_growth(wb)
    extract_efficiency_improvement(wb)
    extract_ocgt_load_factor(wb)
    extract_ocgt_load_shedding(wb)
    extract_passenger_departures(wb)
    extract_refinery_production(wb)
    extract_historical_demand(wb)

    print("\nDone.")


if __name__ == "__main__":
    main()
