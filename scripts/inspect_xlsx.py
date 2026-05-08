"""Read-only inspection of the original xlsx. Throwaway — delete after running."""
from __future__ import annotations

from pathlib import Path

import openpyxl

XLSX = Path("docs/Liquid Fuels Model - Supply Demand, 2025 - Reatile Copy.xlsx")
OUT = Path("_inspect_output.txt")

wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=False)

lines: list[str] = []
add = lines.append

add("=" * 70)
add(f"FILE: {XLSX.name}")
add(f"SIZE: {XLSX.stat().st_size:,} bytes")
add("=" * 70)

add("\n--- SHEETS ---")
for name in wb.sheetnames:
    ws = wb[name]
    add(f"  {name:40s}  dims={ws.dimensions}  rows={ws.max_row}  cols={ws.max_column}")

add("\n--- DEFINED NAMES (named ranges) with resolved values ---")
defined = list(wb.defined_names)
add(f"  total: {len(defined)}\n")

def cell_repr(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v[:50]
    if isinstance(v, float):
        return f"{v:.6g}"
    return str(v)

for nm in defined:
    dn = wb.defined_names[nm]
    add(f"\n  [{nm}]  ref: {dn.value}")
    try:
        for sheet_title, coord in dn.destinations:
            ws = wb[sheet_title]
            cells = ws[coord]
            # cells can be a Cell, a row tuple, or a tuple of rows
            if hasattr(cells, "value"):
                add(f"    -> scalar: {cell_repr(cells.value)}")
            else:
                rows = list(cells)
                # Normalise: if one row, wrap in a list
                if rows and hasattr(rows[0], "value"):
                    rows = [rows]
                add(f"    -> {len(rows)} rows x {len(rows[0]) if rows else 0} cols")
                for ri, row in enumerate(rows[:8]):
                    vals = [cell_repr(c.value) for c in row]
                    add(f"       r{ri+1}: {vals}")
                if len(rows) > 8:
                    add(f"       ... ({len(rows) - 8} more rows)")
    except Exception as e:
        add(f"    ! could not resolve: {e}")

add("\n\n--- KEY SHEET PEEKS ---")
peek_sheets = [
    "Assumptions",
    "Supply Forecast",
    "Jet - DemandSupply",
    "Diesel - DemandSupply",
    "Gasoline - DemandSupply",
    "fEV Penetration",
    "RegEV",
    "RegJetFuel",
    "RegDiesel",
    "RegGasoline",
]
for name in peek_sheets:
    if name not in wb.sheetnames:
        continue
    ws = wb[name]
    add(f"\n[{name}]   ({ws.max_row} rows x {ws.max_column} cols)")
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=20)):
        trimmed = []
        for v in row[:14]:
            if v is None:
                trimmed.append("")
            elif isinstance(v, str):
                trimmed.append(v[:35])
            elif isinstance(v, float):
                trimmed.append(f"{v:.6g}")
            else:
                trimmed.append(str(v))
        add(f"  r{i+1:02d}: {trimmed}")

OUT.write_text("\n".join(lines), encoding="utf-8")
print(f"wrote {len(lines)} lines to {OUT}")
